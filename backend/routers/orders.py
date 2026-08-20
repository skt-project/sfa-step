"""
Unified multi-source orders for Visit & Order.

GET /orders          — merged, filtered, paginated order list + summary + per-source status
GET /orders/detail   — one order's line items (source-aware)
GET /orders/export   — the SAME filtered set as an .xlsx workbook

ADDITIVE BY DESIGN. This router only READS. It does not modify step_visit,
step_visit_item or any other SFA table, and routers/visit.py is untouched and
still serves the handheld and the existing SFA flows exactly as before. The
Google Spreadsheet feed is an ADDITIONAL source alongside SFA, never a
replacement for it.

SOURCE INDEPENDENCE (spec §14): each source is queried in its own try/except.
One failing source yields an error entry in `sources[]` while the other source's
orders are still returned, so a spreadsheet problem can never take Visit & Order
down for SFA — and vice versa.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from dependencies import require_role
from models.auth import UserContext
from models.order import (
    OrderListResponse,
    OrderPagination,
    OrderRow,
    OrderSummary,
    SourceStatus,
)
from services.bq import BQClient
from services.orders import (
    SOURCE_LABELS,
    SOURCE_SFA,
    SOURCE_SHEET,
    SourceUnavailable,
    fetch_sfa_items,
    fetch_sfa_orders,
    fetch_sheet_items,
    fetch_sheet_orders,
    sort_orders,
    summarise,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/orders", tags=["orders"])

# Roles that may open Visit & Order (mirrors the nav entry).
_ALLOWED_ROLES = ("spv", "asm", "dm", "ho_admin")

# Upper bound on rows pulled per source before merging. Sorting and paginating a
# merged list requires both sides in memory; this caps that cost. `truncated` is
# returned so the UI can say so rather than silently showing a partial set.
MAX_MERGE = 2000

WIB = timezone(timedelta(hours=7))


def _filters(
    from_date: str | None, to_date: str | None, status: str | None, store: str | None,
    sku: str | None, order_number: str | None, search: str | None,
) -> dict:
    return {
        "from_date": from_date, "to_date": to_date, "status": status,
        "store": store, "sku": sku, "order_number": order_number, "search": search,
    }


def _collect(
    user: UserContext, source: str, f: dict,
) -> tuple[list[OrderRow], list[SourceStatus]]:
    """Query each requested source independently. A failure is recorded, not raised."""
    bq = BQClient.get()
    wanted = (source or "ALL").upper()
    orders: list[OrderRow] = []
    statuses: list[SourceStatus] = []

    for src, fetch in ((SOURCE_SFA, fetch_sfa_orders), (SOURCE_SHEET, fetch_sheet_orders)):
        if wanted not in ("ALL", src):
            continue
        try:
            rows = fetch(bq, user, f, MAX_MERGE)
            orders.extend(rows)
            statuses.append(SourceStatus(
                source=src, label=SOURCE_LABELS[src], ok=True, count=len(rows),
            ))
        except SourceUnavailable:
            # Generic message: never leak BigQuery internals to the browser.
            statuses.append(SourceStatus(
                source=src, label=SOURCE_LABELS[src], ok=False, count=0,
                error=f"{SOURCE_LABELS[src]} could not be loaded.",
            ))
        except Exception:
            logger.exception("orders: unexpected failure reading %s", src)
            statuses.append(SourceStatus(
                source=src, label=SOURCE_LABELS[src], ok=False, count=0,
                error=f"{SOURCE_LABELS[src]} could not be loaded.",
            ))

    return orders, statuses


@router.get("", response_model=OrderListResponse)
def list_orders(
    from_date: str | None = Query(None, description="Inclusive, YYYY-MM-DD"),
    to_date: str | None = Query(None, description="Inclusive, YYYY-MM-DD"),
    source: str = Query("ALL", description="ALL | SFA | SPREADSHEET"),
    status: str | None = Query(None),
    store: str | None = Query(None, description="Partial store name or store id"),
    sku: str | None = Query(None, description="Partial SKU or product name"),
    order_number: str | None = Query(None),
    search: str | None = Query(None),
    sort_by: str = Query("order_date"),
    sort_order: str = Query("desc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    current_user: UserContext = Depends(require_role(*_ALLOWED_ROLES)),
):
    f = _filters(from_date, to_date, status, store, sku, order_number, search)
    orders, statuses = _collect(current_user, source, f)

    orders = sort_orders(orders, sort_by, sort_order)
    summary = summarise(orders)

    total = len(orders)
    offset = (page - 1) * page_size
    page_rows = orders[offset:offset + page_size]
    total_pages = (total + page_size - 1) // page_size

    return OrderListResponse(
        data=page_rows,
        pagination=OrderPagination(
            page=page, page_size=page_size, total=total,
            total_pages=total_pages, has_next=(offset + page_size) < total,
        ),
        summary=OrderSummary(**summary),
        sources=statuses,
        truncated=any(s.ok and s.count >= MAX_MERGE for s in statuses),
    )


@router.get("/detail")
def order_detail(
    source: str = Query(..., description="SFA | SPREADSHEET"),
    order_id: str = Query(...),
    current_user: UserContext = Depends(require_role(*_ALLOWED_ROLES)),
):
    """Line items for one order.

    Scoping is enforced by re-fetching the order through the SAME filtered query
    the list uses — an id outside the caller's scope simply is not found, so this
    cannot be used to read another distributor's order.
    """
    bq = BQClient.get()
    src = (source or "").upper()
    f = _filters(None, None, None, None, None, order_id, None)

    if src == SOURCE_SFA:
        rows = fetch_sfa_orders(bq, current_user, f, MAX_MERGE)
    elif src == SOURCE_SHEET:
        rows = fetch_sheet_orders(bq, current_user, f, MAX_MERGE)
    else:
        return {"order": None, "items": [], "error": "Unknown source."}

    match = next((r for r in rows if r.order_id == order_id), None)
    if match is None:
        # 404-equivalent without an existence oracle: same shape as "no items".
        return {"order": None, "items": []}

    items = (fetch_sfa_items(bq, [order_id]) if src == SOURCE_SFA
             else fetch_sheet_items(bq, [order_id]))
    return {"order": match, "items": items}


# ---------------------------------------------------------------------------
# Excel export — same filters, same scoping, same rows as the UI
# ---------------------------------------------------------------------------

_SUMMARY_HEADERS = [
    "Source", "Order Number", "Order Date", "Store ID", "Store Name",
    "Distributor Code", "Distributor Name", "Salesman", "Items",
    "Quantity", "Order Value", "Status",
]

_DETAIL_HEADERS = [
    "Source", "Order Number", "Order Date", "Store ID", "Store Name",
    "SKU", "Product Name", "Quantity", "Unit Price", "Line Value",
]


def _autosize(ws) -> None:
    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(ws.columns, start=1):
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 42)


def _write_header(ws, headers: list[str]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A8A")
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


@router.get("/export/{source}/{order_id}")
def export_single_order(
    source: str,
    order_id: str,
    current_user: UserContext = Depends(require_role(*_ALLOWED_ROLES)),
):
    """Export ONE order — same two-sheet shape as the bulk export, scoped exactly
    like order_detail(): re-fetched through the caller's own filtered query, so an
    id outside their scope 404s rather than leaking another distributor's order."""
    from openpyxl import Workbook

    bq = BQClient.get()
    src = (source or "").upper()
    f = _filters(None, None, None, None, None, order_id, None)

    if src == SOURCE_SFA:
        rows = fetch_sfa_orders(bq, current_user, f, MAX_MERGE)
    elif src == SOURCE_SHEET:
        rows = fetch_sheet_orders(bq, current_user, f, MAX_MERGE)
    else:
        raise HTTPException(status_code=404, detail="Not found")

    order = next((r for r in rows if r.order_id == order_id), None)
    if order is None:
        raise HTTPException(status_code=404, detail="Not found")

    items = (fetch_sfa_items(bq, [order_id]) if src == SOURCE_SFA
             else fetch_sheet_items(bq, [order_id]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Order Summary"
    _write_header(ws, _SUMMARY_HEADERS)
    ws.append([
        order.source_label, order.order_number, order.order_date, order.store_id,
        order.store_name, order.distributor_code, order.distributor_name,
        order.salesman_name, order.item_count, order.quantity, order.order_value,
        order.status,
    ])
    for row in ws.iter_rows(min_row=2, min_col=10, max_col=11):
        for cell in row:
            cell.number_format = "#,##0.##"
    _autosize(ws)

    ws2 = wb.create_sheet("Order Details")
    _write_header(ws2, _DETAIL_HEADERS)
    for it in items:
        ws2.append([
            order.source_label, it.order_number, order.order_date, order.store_id,
            order.store_name, it.sku, it.product_name, it.quantity, it.unit_price,
            it.line_value,
        ])
    for row in ws2.iter_rows(min_row=2, min_col=8, max_col=10):
        for cell in row:
            cell.number_format = "#,##0.##"
    _autosize(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Order_{order.order_id}.xlsx"
    logger.info("orders/export-single: user=%s source=%s order_id=%s",
                current_user.username, source, order_id)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export")
def export_orders(
    from_date: str | None = Query(None),
    to_date: str | None = Query(None),
    source: str = Query("ALL"),
    status: str | None = Query(None),
    store: str | None = Query(None),
    sku: str | None = Query(None),
    order_number: str | None = Query(None),
    search: str | None = Query(None),
    sort_by: str = Query("order_date"),
    sort_order: str = Query("desc"),
    current_user: UserContext = Depends(require_role(*_ALLOWED_ROLES)),
):
    """Export exactly what the UI is showing: every active filter and the
    caller's scoping are applied identically, so the workbook can never contain a
    row the user could not see on screen. Two sheets — one row per ORDER, then
    one row per line ITEM (SKU/product live at item level)."""
    from openpyxl import Workbook

    bq = BQClient.get()
    f = _filters(from_date, to_date, status, store, sku, order_number, search)
    orders, statuses = _collect(current_user, source, f)
    orders = sort_orders(orders, sort_by, sort_order)

    wb = Workbook()

    # ── Sheet 1: one row per order ──
    ws = wb.active
    ws.title = "Order Summary"
    _write_header(ws, _SUMMARY_HEADERS)
    for o in orders:
        ws.append([
            o.source_label, o.order_number, o.order_date, o.store_id, o.store_name,
            o.distributor_code, o.distributor_name, o.salesman_name, o.item_count,
            o.quantity, o.order_value, o.status,
        ])
    for row in ws.iter_rows(min_row=2, min_col=10, max_col=11):
        for cell in row:
            cell.number_format = "#,##0.##"
    _autosize(ws)

    # ── Sheet 2: one row per line item ──
    ws2 = wb.create_sheet("Order Details")
    _write_header(ws2, _DETAIL_HEADERS)
    by_id = {o.order_id: o for o in orders}
    sfa_ids = [o.order_id for o in orders if o.source == SOURCE_SFA]
    sheet_ids = [o.order_id for o in orders if o.source == SOURCE_SHEET]
    items = []
    for ids, fetch in ((sfa_ids, fetch_sfa_items), (sheet_ids, fetch_sheet_items)):
        for start in range(0, len(ids), 400):        # keep the IN() list sane
            try:
                items.extend(fetch(bq, ids[start:start + 400]))
            except Exception:
                logger.exception("orders/export: item fetch failed; continuing")
    for it in items:
        parent = by_id.get(it.order_id)
        ws2.append([
            SOURCE_LABELS.get(it.source, it.source),
            it.order_number,
            parent.order_date if parent else None,
            parent.store_id if parent else None,
            parent.store_name if parent else None,
            it.sku, it.product_name, it.quantity, it.unit_price, it.line_value,
        ])
    for row in ws2.iter_rows(min_row=2, min_col=8, max_col=10):
        for cell in row:
            cell.number_format = "#,##0.##"
    _autosize(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    stamp = datetime.now(WIB).strftime("%Y%m%d_%H%M")
    src_tag = {"ALL": "AllSources", SOURCE_SFA: "SFA", SOURCE_SHEET: "Spreadsheet"}.get(
        (source or "ALL").upper(), "Orders"
    )
    filename = f"VisitOrder_{src_tag}_{stamp}.xlsx"
    logger.info(
        "orders/export: user=%s source=%s orders=%s items=%s",
        current_user.username, source, len(orders), len(items),
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
